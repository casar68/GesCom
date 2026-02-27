import io
from decimal import Decimal

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from app.models.facture import Facture


def generate_facture_pdf(facture: Facture, client_name: str) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()
    elements = []

    # Titre
    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=18, spaceAfter=12)
    elements.append(Paragraph(f"FACTURE {facture.numero}", title_style))
    elements.append(Spacer(1, 10 * mm))

    # Infos
    info_style = styles["Normal"]
    elements.append(Paragraph(f"<b>Client :</b> {client_name}", info_style))
    elements.append(Paragraph(f"<b>Date :</b> {facture.date_facture.strftime('%d/%m/%Y')}", info_style))
    if facture.date_echeance:
        elements.append(Paragraph(f"<b>Echéance :</b> {facture.date_echeance.strftime('%d/%m/%Y')}", info_style))
    if facture.mode_reglement:
        elements.append(Paragraph(f"<b>Règlement :</b> {facture.mode_reglement}", info_style))
    elements.append(Spacer(1, 10 * mm))

    # Lignes
    table_data = [["#", "Désignation", "Qté", "PU HT", "Remise", "Montant HT"]]
    for ligne in facture.lignes:
        table_data.append([
            str(ligne.ligne_numero),
            ligne.designation,
            str(ligne.quantite),
            f"{ligne.prix_unitaire_ht:.2f} €",
            f"{ligne.remise_pct:.1f}%" if ligne.remise_pct else "-",
            f"{ligne.montant_ht:.2f} €",
        ])

    table = Table(table_data, colWidths=[15 * mm, 80 * mm, 15 * mm, 25 * mm, 20 * mm, 30 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 10 * mm))

    # Totaux
    totaux_data = [
        ["Total HT", f"{facture.total_ht:.2f} €"],
        ["TVA", f"{facture.total_tva:.2f} €"],
        ["Total TTC", f"{facture.total_ttc:.2f} €"],
    ]
    if facture.montant_regle > 0:
        totaux_data.append(["Déjà réglé", f"{facture.montant_regle:.2f} €"])
        reste = facture.total_ttc - facture.montant_regle
        totaux_data.append(["Reste à payer", f"{reste:.2f} €"])

    totaux_table = Table(totaux_data, colWidths=[120 * mm, 40 * mm])
    totaux_table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("LINEABOVE", (0, -1), (-1, -1), 1, colors.black),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    elements.append(totaux_table)

    doc.build(elements)
    return buffer.getvalue()


def generate_excel_report(data: list[dict], sheet_name: str = "Rapport") -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not data:
        ws.append(["Aucune donnée"])
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # En-têtes
    headers = list(data[0].keys())
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Données
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(key))

    # Auto-width
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
